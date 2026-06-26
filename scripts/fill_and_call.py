#!/usr/bin/env python3
"""
fill_and_call.py — qwen-question-gpt-audit 的通用「填变量 + 调用 qwen-plus」脚本。

职责（仅生成，不审核）：
- 读取 Excel，自动识别三种结构：
  1) 含「出题」「出答案」两个 sheet 的重出题文件 → 用「出题」提示词出题，再用「出答案」提示词出该题答案；
  2) 单 sheet 的重出答案文件 → 保持题干不变，仅重出三层答案；
  3) 通用单 sheet 出题文件 → 每行用「提示词」列出题。
- 按变量映射表把模板里的 {{占位符}} 原样替换为该行字段值（模板逐字取用，不改写/缩写/扩展）。
- 调用 DashScope OpenAI 兼容接口的 qwen-plus，落 JSON 供当前 GPT agent 逐字按 prompts/审题提示词.txt 审核。

注意：本脚本不做审核、不判定「通过」。审核必须由 GPT 按 prompts/审题提示词.txt 逐字完成。
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from typing import Any
from urllib import request, error

import openpyxl

API_URL = "https://dashscope.aliyuncs.com/compatible-mode/v1/chat/completions"
MODEL = "qwen-plus"

# 变量映射：模板占位符 -> Excel 候选列名（按顺序取第一个存在的）
VAR_TO_COLUMNS: dict[str, list[str]] = {
    "industry": ["行业"],
    "job": ["岗位", "岗位名称"],
    "scene": ["招聘场景", "场景"],
    "category": ["技能分类"],
    "skill": ["技能"],
    "knowledge": ["知识点"],
    "level": ["难度"],
    "JD": ["JD"],
    "Q": ["题目", "问题"],
}

QUESTION_COL_CANDIDATES = ["题目"]
PROMPT_COL = "提示词"

OUTPUT_COLUMNS = [
    "招聘场景", "行业", "岗位", "JD", "技能分类", "技能", "知识点", "难度", "题目",
    "参考答案-第一层", "参考答案-第二层", "参考答案-第三层",
]


def normalize_template(text: Any) -> str:
    t = "" if text is None else str(text)
    t = t.strip()
    if len(t) >= 2 and t[0] == '"' and t[-1] == '"':
        t = t[1:-1]
    # 兼容模板中 {{{industry}} 这类缺括号写法
    t = t.replace("{{{industry}}", "{{industry}}")
    return t


def build_mapping(row: dict[str, Any], *, num: int = 1, question: str | None = None) -> dict[str, str]:
    mapping: dict[str, str] = {"num": str(num)}
    for var, cols in VAR_TO_COLUMNS.items():
        val = ""
        for c in cols:
            if c in row and row[c] is not None and str(row[c]).strip():
                val = str(row[c])
                break
        mapping[var] = val
    if question is not None:
        mapping["Q"] = question
    return mapping


def replace_vars(template: Any, mapping: dict[str, str]) -> str:
    out = normalize_template(template)
    for key, value in mapping.items():
        out = out.replace("{{" + key + "}}", value)
    return out


def qwen_chat(prompt: str, *, temperature: float = 0.4, extra: str = "") -> str:
    api_key = os.environ.get("DASHSCOPE_API_KEY")
    if not api_key:
        raise RuntimeError("环境变量 DASHSCOPE_API_KEY 未设置")
    payload = {
        "model": MODEL,
        "messages": [
            {"role": "system", "content": "你只按用户要求输出严格 JSON，不输出额外解释。"},
            {"role": "user", "content": prompt + extra},
        ],
        "temperature": temperature,
        "response_format": {"type": "json_object"},
    }
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = request.Request(
        API_URL, data=data,
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        method="POST",
    )
    last_err: Exception | None = None
    for attempt in range(5):
        try:
            with request.urlopen(req, timeout=120) as resp:
                obj = json.loads(resp.read().decode("utf-8"))
            return obj["choices"][0]["message"]["content"]
        except error.HTTPError as exc:
            if exc.code < 500:
                body = exc.read().decode("utf-8", errors="ignore")
                raise RuntimeError(f"qwen-plus HTTP {exc.code}: {body[:300]}") from exc
            last_err = exc
        except (error.URLError, TimeoutError) as exc:
            last_err = exc
        time.sleep(3 * (attempt + 1))
    raise RuntimeError(f"qwen-plus 多次请求失败：{last_err}")


def parse_json_content(text: str) -> dict[str, Any]:
    cleaned = text.strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?", "", cleaned).strip()
        cleaned = re.sub(r"```$", "", cleaned).strip()
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        match = re.search(r"\{[\s\S]*\}", cleaned)
        if not match:
            raise
        return json.loads(match.group(0))


def first_item(obj: dict[str, Any]) -> dict[str, Any]:
    qs = obj.get("questions")
    if isinstance(qs, list) and qs:
        return qs[0]
    return obj


def sheet_to_rows(ws) -> list[dict[str, Any]]:
    header = [c.value for c in ws[1]]
    rows: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        row = {header[i]: ws.cell(r, i + 1).value for i in range(len(header))}
        if any(v is not None and str(v).strip() for v in row.values()):
            rows.append(row)
    return rows


def find_question_col(row: dict[str, Any]) -> str | None:
    for c in QUESTION_COL_CANDIDATES:
        if c in row:
            return c
    return None


def base_record(row: dict[str, Any]) -> dict[str, Any]:
    def g(*names: str) -> Any:
        for n in names:
            if n in row and row[n] is not None:
                return row[n]
        return ""
    return {
        "招聘场景": g("招聘场景", "场景"),
        "行业": g("行业"),
        "岗位": g("岗位", "岗位名称"),
        "JD": g("JD"),
        "技能分类": g("技能分类"),
        "技能": g("技能"),
        "知识点": g("知识点"),
        "难度": g("难度"),
    }


def generate_from_question_sheet(wb, *, extra_q: str, extra_a: str) -> list[dict[str, Any]]:
    """结构1：含「出题」「出答案」两个 sheet。出题 sheet 每行出 1 题，再用 出答案 sheet 提示词出答案。"""
    q_ws = wb["出题"]
    a_ws = wb["出答案"]
    answer_tpl = a_ws.cell(2, 2).value if a_ws.max_column >= 2 else a_ws.cell(2, 1).value
    results: list[dict[str, Any]] = []
    for row in sheet_to_rows(q_ws):
        q_tpl = row.get(PROMPT_COL)
        qmap = build_mapping(row, num=1)
        q_out = first_item(parse_json_content(qwen_chat(replace_vars(q_tpl, qmap), extra=extra_q)))
        question = str(q_out.get("question", "")).strip()
        amap = build_mapping(row, question=question)
        a_out = first_item(parse_json_content(qwen_chat(replace_vars(answer_tpl, amap), extra=extra_a)))
        rec = base_record(row)
        rec.update({
            "题目": question,
            "level": q_out.get("level"),
            "time": q_out.get("time"),
            "参考答案-第一层": str(a_out.get("answer1", "")).strip(),
            "参考答案-第二层": str(a_out.get("answer2", "")).strip(),
            "参考答案-第三层": str(a_out.get("answer3", "")).strip(),
            "_source": "出题",
        })
        results.append(rec)
    return results


def generate_answers_only(wb, *, fallback_answer_tpl: Any, extra_a: str) -> list[dict[str, Any]]:
    """结构2：单 sheet，保持题干，仅重出三层答案。提示词取行内『提示词』列，空则用 fallback。"""
    ws = wb[wb.sheetnames[0]]
    results: list[dict[str, Any]] = []
    for row in sheet_to_rows(ws):
        qcol = find_question_col(row)
        question = str(row.get(qcol, "")).strip() if qcol else ""
        row_tpl = row.get(PROMPT_COL)
        answer_tpl = row_tpl if (row_tpl and str(row_tpl).strip()) else fallback_answer_tpl
        if not (answer_tpl and str(answer_tpl).strip()):
            raise RuntimeError("缺少出答案提示词：行内『提示词』为空且未提供 --answer-prompt-file")
        amap = build_mapping(row, question=question)
        a_out = first_item(parse_json_content(qwen_chat(replace_vars(answer_tpl, amap), extra=extra_a)))
        rec = base_record(row)
        rec.update({
            "题目": question,
            "参考答案-第一层": str(a_out.get("answer1", "")).strip(),
            "参考答案-第二层": str(a_out.get("answer2", "")).strip(),
            "参考答案-第三层": str(a_out.get("answer3", "")).strip(),
            "_source": "出答案",
        })
        results.append(rec)
    return results


def generate_questions_only(wb, *, extra_q: str) -> list[dict[str, Any]]:
    """结构3：通用单 sheet 出题文件，每行用『提示词』列出题（不出答案）。"""
    ws = wb[wb.sheetnames[0]]
    results: list[dict[str, Any]] = []
    for row in sheet_to_rows(ws):
        q_tpl = row.get(PROMPT_COL)
        if not (q_tpl and str(q_tpl).strip()):
            raise RuntimeError("通用出题文件每行需有『提示词』列")
        qmap = build_mapping(row, num=1)
        q_out = first_item(parse_json_content(qwen_chat(replace_vars(q_tpl, qmap), extra=extra_q)))
        rec = base_record(row)
        rec.update({
            "题目": str(q_out.get("question", "")).strip(),
            "level": q_out.get("level"),
            "time": q_out.get("time"),
            "_source": "出题(单sheet)",
        })
        results.append(rec)
    return results


def detect_and_run(path: str, *, mode: str, fallback_tpl_file: str | None,
                   extra_q: str, extra_a: str) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path)
    fallback_tpl = None
    if fallback_tpl_file:
        with open(fallback_tpl_file, encoding="utf-8") as f:
            fallback_tpl = f.read()

    if mode == "auto":
        if "出题" in wb.sheetnames and "出答案" in wb.sheetnames:
            mode = "question"
        else:
            ws = wb[wb.sheetnames[0]]
            header = [c.value for c in ws[1]]
            mode = "answers" if any(h in QUESTION_COL_CANDIDATES for h in header) else "questions"

    if mode == "question":
        return generate_from_question_sheet(wb, extra_q=extra_q, extra_a=extra_a)
    if mode == "answers":
        return generate_answers_only(wb, fallback_answer_tpl=fallback_tpl, extra_a=extra_a)
    if mode == "questions":
        return generate_questions_only(wb, extra_q=extra_q)
    raise ValueError(f"未知 mode: {mode}")


def main() -> None:
    ap = argparse.ArgumentParser(description="qwen-plus 填变量并生成（仅生成，不审核）")
    ap.add_argument("excel", help="输入 Excel 路径")
    ap.add_argument("-o", "--out", required=True, help="输出 JSON 路径（供 GPT 审核）")
    ap.add_argument("--mode", default="auto",
                    choices=["auto", "question", "answers", "questions"],
                    help="auto 自动识别；question=两sheet出题+出答案；answers=保留题干重出答案；questions=单sheet仅出题")
    ap.add_argument("--answer-prompt-file", default=None,
                    help="结构2 中行内『提示词』为空时使用的出答案提示词模板文件")
    ap.add_argument("--extra-q-file", default=None, help="出题阶段追加约束文件（重出迭代用）")
    ap.add_argument("--extra-a-file", default=None, help="出答案阶段追加约束文件（重出迭代用）")
    args = ap.parse_args()

    def read_opt(p: str | None) -> str:
        if not p:
            return ""
        with open(p, encoding="utf-8") as f:
            return "\n\n" + f.read().strip()

    extra_q = read_opt(args.extra_q_file)
    extra_a = read_opt(args.extra_a_file)

    results = detect_and_run(args.excel, mode=args.mode,
                             fallback_tpl_file=args.answer_prompt_file,
                             extra_q=extra_q, extra_a=extra_a)
    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)

    print(f"generated={len(results)}  out={args.out}")
    print("提醒：结果尚未审核。请由 GPT agent 逐字按 prompts/审题提示词.txt 审核后再落最终 Excel。")


if __name__ == "__main__":
    main()
